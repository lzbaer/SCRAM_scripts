import pandas as pd
import json
import math
from collections import OrderedDict
from openpyxl import load_workbook
import xlsxwriter

MAPPED_CIS_OUTPUT = "cis_from_mitre"

ATTACK_JSON = "../aggregate.json"

CIS_MAPPINGS_FILE = "CIS_Controls_v8_to_Enterprise_ATTCK_v82_Master_Mapping__5262021.xlsx"
CIS_MAPPINGS_SHEET = "V8-ATT&CK Low Mit. & (Sub-)Tech"
CIS_CONTROL_COL = "CIS Control"
MIT_COL = "ATT&CK V8.2 Enterprise Mitigation ID"
TECH_COL = "ATT&CK Technique ID"
SUBTECH_COL = "Combined ATT&CK (Sub-)Technique ID"
CIS_NAME_COL = "Title"

def main():

	# Read the Excel file into a DataFrame
	df = pd.read_excel(CIS_MAPPINGS_FILE, sheet_name=CIS_MAPPINGS_SHEET)

	tech_mappings = {}
	mit_mappings = {}

	cis_counts_t = {}
	cis_counts_m = {}

	for index, row in df.iterrows():
		tech, mit, cis_id, cis_name = row[SUBTECH_COL], row[MIT_COL], row[CIS_CONTROL_COL], row[CIS_NAME_COL]
		if pd.isna(tech) and pd.isna(mit):
			continue

		if mit not in mit_mappings:
			mit_mappings[mit] = [cis_name]
		else:
			mit_mappings[mit].append(cis_name)

		if tech not in tech_mappings:
			tech_mappings[tech] = [cis_name]
		else:
			tech_mappings[tech].append(cis_name)

		if cis_name not in cis_counts_m:
			cis_counts_m[cis_name] = {"id": cis_id, "count": 0}

		if cis_name not in cis_counts_t:
			cis_counts_t[cis_name] = {"id": cis_id, "count": 0}



	with open(ATTACK_JSON, "r") as file:
		attack_data = json.load(file)

	unknown_t = {}
	unknown_m = {}
	
	for name, vals in attack_data["Techniques"].items():
		if vals["id"] not in tech_mappings:
			unknown_t[name] = {"id": vals["id"], "count": vals["count"]}
			# print("{}:{} an unrecognized technique (not in MITRE ATT&CK version 8?)".format(vals["id"], name))
		else:
			for cis in tech_mappings[vals["id"]]:
				cis_counts_t[cis]["count"] += vals["count"]

	print("Unknown Techniques: {}".format(len(unknown_t)))

	for name, vals in attack_data["Mitigations"].items():
		if vals["id"] not in mit_mappings:
			unknown_m[name] = {"id": vals["id"], "count": vals["count"]}
		else:
			for cis in mit_mappings[vals["id"]]:
				cis_counts_m[cis]["count"] += vals["count"]

	print("Unknown Mitigations: {}".format(len(unknown_m)))

	data = {}
	for label, dataset in [("Mitigations", cis_counts_m), ("Techniques", cis_counts_t), ("Excluded Mitigations", unknown_m), ("Excluded Techniques", unknown_t)]:
		sortd = OrderedDict(sorted(dataset.items(), key=lambda item: item[1]["count"], reverse=True))
		data[label] = sortd


	with open("{}.json".format(MAPPED_CIS_OUTPUT),  "w") as file:
		json.dump(data, file, indent=4)

	# to excel
	efile = '{}.xlsx'.format(MAPPED_CIS_OUTPUT)

	# Create a workbook and add a worksheet.
	workbook = xlsxwriter.Workbook(efile)
	

	for label in data.keys():
		workbook.add_worksheet(label)

	workbook.close()



	workbook = load_workbook(efile)
	writer = pd.ExcelWriter(efile, engine='openpyxl')
	# writer.book = workbook
	# writer.sheets = {ws.title: ws for ws in workbook.worksheets}

	for (sheet, dataset) in data.items():
		df = pd.DataFrame(dataset)
		df = df.T
		df.to_excel(writer, sheet_name=sheet)
		# df.to_excel(writer, startrow=writer.sheets[sheet].max_row, index = False, header= False)
		# df.to_excel(, sheet_name=sheet, index=False)
	
	writer.close()

	



main()